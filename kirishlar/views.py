from django.shortcuts import render
from rest_framework.generics import ListCreateAPIView, RetrieveUpdateDestroyAPIView, ListAPIView
from rest_framework.views import APIView
from datetime import datetime, timedelta
from rest_framework.response import Response
from kirishlar.models import Kirish
from kirishlar.serializer import KirishSerializer
from oquvchi.models import Oquvchi
from datetime import datetime, timedelta
from django.utils import timezone
from rest_framework.views import APIView
from rest_framework.response import Response

import os
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
# from openpyxl.writer.excel import save_virtual_workbook

class ExcelReportAPIView(APIView):
    def get(self, request, maktab, kun):
        days = int(kun)
        start_date = timezone.now() - timedelta(days=days)
        wb = Workbook()
        ws = wb.active
        
        ws["A1"] = "Maktab"
        ws["B1"] = "Sinf"
        ws["C1"] = "Ism-familya"
        ws["D1"] = "Telefon raqam"
        ws["E1"] = "Manzil"
        ws["F1"] = "Sana"
        ws["G1"] = "Vaqt"
        i = 2
        for row in Kirish.objects.values_list(): 
                print( row[2], datetime.now().date() - timedelta(days=days),  row[2] >= datetime.now().date() - timedelta(days=days))
                oquvchi = Oquvchi.objects.get(pk=row[1]) 
                if oquvchi.maktab == maktab and row[2] >= datetime.now().date() - timedelta(days=days):
                    ws[f"A{i}"] = oquvchi.maktab
                    ws[f"B{i}"] = oquvchi.sinf
                    ws[f"C{i}"] = oquvchi.ism_familya
                    ws[f"D{i}"] = oquvchi.phone
                    ws[f"E{i}"] = oquvchi.manzil
                    ws[f"F{i}"] = row[2]
                    ws[f"G{i}"] = row[3]
                    i += 1
        
        documents_path = 'report.xlsx'

        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[column_letter].width = adjusted_width


        wb.save(documents_path)

        with open(documents_path, 'rb') as excel_file:
            data = excel_file.read()

        response = HttpResponse(data, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="report.xlsx"'
        return response




# Create your views here.
class AddOrGet(ListCreateAPIView):
    queryset = Kirish.objects.all()
    serializer_class = KirishSerializer

class UpdateOrDel(RetrieveUpdateDestroyAPIView):
    queryset = Kirish.objects.all()
    serializer_class = KirishSerializer

class GetOquvchi(ListAPIView):
    queryset = Kirish.objects.all()
    serializer_class = KirishSerializer

    lookup_field = "oquvchi_id"

class GetSana(RetrieveUpdateDestroyAPIView):
    queryset = Kirish.objects.all()
    serializer_class = KirishSerializer

class GetSana(APIView):
    def get(self, request, kun):
        days = int(kun)
        start_date = timezone.now() - timedelta(days=days)
        records = Kirish.objects.filter(sana__gte=start_date)
        
        serializer = KirishSerializer(records, many=True)
        return Response(serializer.data)
